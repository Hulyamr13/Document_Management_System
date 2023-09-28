import os
import psycopg2
from docx import Document
from openpyxl import load_workbook
import PyPDF2

# Function to create the documents table
def create_documents_table():
    conn = psycopg2.connect(
        host='localhost',
        database='document_manager',
        user='postgres-user',
        password='password'
    )
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id SERIAL PRIMARY KEY,
            filename VARCHAR(255),
            file_type VARCHAR(10),
            content TEXT
        );
    ''')
    conn.commit()
    conn.close()

# Function to insert a document into the database
def insert_document(filename, file_type, content):
    conn = psycopg2.connect(
        host='localhost',
        database='document_manager',
        user='postgres-user',
        password='password'
    )
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO documents (filename, file_type, content) VALUES (%s, %s, %s)",
        (filename, file_type, content)
    )
    conn.commit()
    conn.close()

# Directory where you will find the documents
documents_directory = "C:/Documents"

# Create the documents table
create_documents_table()

# Walk through all files in the directory
for root, dirs, files in os.walk(documents_directory):
    for filename in files:
        file_path = os.path.join(root, filename)
        file_extension = filename.split('.')[-1].lower()

        if file_extension in ('docx', 'xlsx', 'pdf'):
            # Read and process the content of the document
            content = ""
            if file_extension == 'docx':
                doc = Document(file_path)
                for para in doc.paragraphs:
                    content += para.text + '\n'
            elif file_extension == 'xlsx':
                wb = load_workbook(file_path)
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True):
                        content += ' '.join([str(cell) for cell in row]) + '\n'
            elif file_extension == 'pdf':
                pdf_file = open(file_path, 'rb')
                pdf_reader = PyPDF2.PdfFileReader(pdf_file)
                for page_num in range(pdf_reader.numPages):
                    page = pdf_reader.getPage(page_num)
                    content += page.extractText() + '\n'

            # Save the document in the database
            insert_document(filename, file_extension, content)
            print(f"Saved: {file_path}")

print("All documents processed and saved in the database.")
